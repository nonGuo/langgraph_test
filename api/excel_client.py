"""
Excel 生成客户端 - 本地实现.

该模块使用 openpyxl 库在本地生成 Excel 文件，无需依赖外部 API.

功能特性:
- 根据测试用例数据生成格式化的 Excel 文件
- 支持自定义表头映射
- 自动调整列宽
- 支持交替行背景色
- 支持保存为文件或返回字节流
"""

import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Side,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from excel_config import (
    TEST_CASE_FIELD_HEADERS,
    TEST_CASE_FIELD_ORDER,
    TEST_CASE_COLUMN_WIDTHS,
    ExcelStyleConfig,
    DEFAULT_EXCEL_FILENAME_PREFIX,
    FILENAME_DATE_FORMAT,
    get_field_header,
    get_ordered_fields,
    get_column_width,
    convert_field_value,
)

logger = logging.getLogger(__name__)


@dataclass
class ExcelGenerationResult:
    """Excel 生成结果."""
    success: bool
    file_path: Optional[str] = None      # 本地文件路径
    file_url: Optional[str] = None       # 文件 URL（兼容旧接口）
    file_content: Optional[bytes] = None # 二进制文件内容
    error: Optional[str] = None
    row_count: int = 0                   # 生成的数据行数


class ExcelClient:
    """
    本地 Excel 生成客户端.

    使用 openpyxl 库在本地生成 Excel 文件，无需依赖外部 API.

    Attributes:
        output_dir: 输出目录，默认为 ./excel_output
        filename_prefix: 文件名前缀
        default_sheet_name: 默认工作表名称
    """

    def __init__(
        self,
        output_dir: str = "./excel_output",
        filename_prefix: str = DEFAULT_EXCEL_FILENAME_PREFIX,
        default_sheet_name: str = "测试用例",
    ):
        self.output_dir = output_dir
        self.filename_prefix = filename_prefix
        self.default_sheet_name = default_sheet_name

        # 确保输出目录存在
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_excel(
        self,
        test_cases: list[dict[str, Any]],
        filename: Optional[str] = None,
        return_content: bool = False,
    ) -> ExcelGenerationResult:
        """
        生成 Excel 文件.

        Args:
            test_cases: 测试用例列表，每个元素为字典
            filename: 输出文件名，如不提供则自动生成
            return_content: 是否返回文件内容字节流

        Returns:
            ExcelGenerationResult 生成结果
        """
        logger.info(f"开始生成 Excel，共 {len(test_cases)} 个测试用例")

        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = self.default_sheet_name

            # 获取字段顺序
            fields = get_ordered_fields()

            # 生成表头
            self._write_headers(ws, fields)

            # 生成数据行
            row_count = self._write_data_rows(ws, test_cases, fields)

            # 应用样式
            self._apply_styles(ws, len(fields), row_count + 1)

            # 调整列宽
            self._adjust_column_widths(ws, fields, test_cases)

            # 生成文件名
            if filename is None:
                filename = self._generate_filename()

            # 确保文件名有正确的后缀
            if not filename.endswith(".xlsx"):
                filename = f"{filename}.xlsx"

            # 保存文件
            file_path = os.path.join(self.output_dir, filename)
            wb.save(file_path)

            logger.info(f"Excel 文件已保存：{file_path}, 共 {row_count} 行数据")

            result = ExcelGenerationResult(
                success=True,
                file_path=file_path,
                file_url=file_path,  # 兼容旧接口
                row_count=row_count,
            )

            # 如果需要返回文件内容
            if return_content:
                result.file_content = self._get_file_content(file_path)

            return result

        except Exception as e:
            logger.exception("Excel 生成失败")
            return ExcelGenerationResult(
                success=False,
                error=f"Excel 生成失败：{str(e)}",
            )

    def generate_excel_sync(
        self,
        test_cases: list[dict[str, Any]],
        filename: Optional[str] = None,
    ) -> ExcelGenerationResult:
        """
        同步版本（兼容旧接口）.

        Args:
            test_cases: 测试用例列表
            filename: 输出文件名

        Returns:
            ExcelGenerationResult 生成结果
        """
        return self.generate_excel(test_cases, filename)

    async def generate_excel_async(
        self,
        test_cases: list[dict[str, Any]],
        filename: Optional[str] = None,
    ) -> ExcelGenerationResult:
        """
        异步版本（兼容旧接口）.

        Args:
            test_cases: 测试用例列表
            filename: 输出文件名

        Returns:
            ExcelGenerationResult 生成结果
        """
        # 由于是本地 IO，直接使用同步版本
        return self.generate_excel(test_cases, filename)

    def _write_headers(
        self,
        ws: Any,
        fields: list[str],
    ) -> None:
        """
        写入表头行.

        Args:
            ws: 工作表对象
            fields: 字段列表
        """
        for col_idx, field in enumerate(fields, start=1):
            header_text = get_field_header(field)
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.value = header_text

    def _write_data_rows(
        self,
        ws: Any,
        test_cases: list[dict[str, Any]],
        fields: list[str],
    ) -> int:
        """
        写入数据行.

        Args:
            ws: 工作表对象
            test_cases: 测试用例列表
            fields: 字段列表

        Returns:
            写入的行数
        """
        for row_idx, test_case in enumerate(test_cases, start=2):
            for col_idx, field in enumerate(fields, start=1):
                # 获取原始值
                raw_value = test_case.get(field, "")

                # 应用转换器
                value = convert_field_value(field, raw_value)

                # 处理 None 值
                if value is None:
                    value = ""

                # 写入单元格
                ws.cell(row=row_idx, column=col_idx, value=value)

        return len(test_cases)

    def _apply_styles(
        self,
        ws: Any,
        num_cols: int,
        num_rows: int,
    ) -> None:
        """
        应用样式（表头样式、边框、交替行背景色）.

        Args:
            ws: 工作表对象
            num_cols: 列数
            num_rows: 行数
        """
        # 定义边框
        thin_border = Border(
            left=Side(style=ExcelStyleConfig.BORDER_STYLE, color=ExcelStyleConfig.BORDER_COLOR),
            right=Side(style=ExcelStyleConfig.BORDER_STYLE, color=ExcelStyleConfig.BORDER_COLOR),
            top=Side(style=ExcelStyleConfig.BORDER_STYLE, color=ExcelStyleConfig.BORDER_COLOR),
            bottom=Side(style=ExcelStyleConfig.BORDER_STYLE, color=ExcelStyleConfig.BORDER_COLOR),
        )

        # 表头样式
        header_font = Font(
            name=ExcelStyleConfig.FONT_NAME,
            size=ExcelStyleConfig.FONT_SIZE,
            bold=ExcelStyleConfig.HEADER_FONT_BOLD,
            color=ExcelStyleConfig.HEADER_FONT_COLOR,
        )
        header_fill = PatternFill(
            start_color=ExcelStyleConfig.HEADER_BG_COLOR,
            end_color=ExcelStyleConfig.HEADER_BG_COLOR,
            fill_type="solid",
        )
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        # 数据行样式
        data_font = Font(
            name=ExcelStyleConfig.FONT_NAME,
            size=ExcelStyleConfig.FONT_SIZE,
            color=ExcelStyleConfig.ROW_FONT_COLOR,
        )
        data_alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        alt_row_fill = PatternFill(
            start_color=ExcelStyleConfig.ROW_BG_COLOR_ALT,
            end_color=ExcelStyleConfig.ROW_BG_COLOR_ALT,
            fill_type="solid",
        )

        # 应用表头样式
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.row_dimensions[1].height = ExcelStyleConfig.HEADER_ROW_HEIGHT

        # 应用数据行样式
        for row in range(2, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

                # 交替行背景色
                if row % 2 == 0:
                    cell.fill = alt_row_fill

            ws.row_dimensions[row].height = ExcelStyleConfig.DEFAULT_ROW_HEIGHT

    def _adjust_column_widths(
        self,
        ws: Any,
        fields: list[str],
        test_cases: list[dict[str, Any]],
    ) -> None:
        """
        调整列宽.

        Args:
            ws: 工作表对象
            fields: 字段列表
            test_cases: 测试用例列表
        """
        for col_idx, field in enumerate(fields, start=1):
            # 获取配置的列宽
            configured_width = get_column_width(field)

            # 获取表头长度
            header_length = len(get_field_header(field))

            # 计算数据列的最大长度
            max_data_length = 0
            for test_case in test_cases:
                value = test_case.get(field, "")
                if value is not None:
                    value_str = str(value)
                    # 中文字符算 2 个宽度
                    chinese_chars = sum(1 for c in value_str if '\u4e00' <= c <= '\u9fff')
                    length = len(value_str) + chinese_chars
                    max_data_length = max(max_data_length, length)

            # 计算最终列宽（取配置值、表头长度、数据最大长度的最大值）
            final_width = max(
                configured_width,
                header_length + ExcelStyleConfig.AUTO_WIDTH_PADDING,
                min(max_data_length + ExcelStyleConfig.AUTO_WIDTH_PADDING, 80),  # 最大 80
            )

            # 设置列宽
            ws.column_dimensions[get_column_letter(col_idx)].width = final_width

    def _generate_filename(self) -> str:
        """
        生成默认文件名.

        Returns:
            文件名（不含路径）
        """
        timestamp = datetime.now().strftime(FILENAME_DATE_FORMAT)
        return f"{self.filename_prefix}{timestamp}.xlsx"

    def _get_file_content(self, file_path: str) -> bytes:
        """
        读取文件内容.

        Args:
            file_path: 文件路径

        Returns:
            文件字节内容
        """
        with open(file_path, "rb") as f:
            return f.read()


# ============================================================================
# 便捷函数
# ============================================================================

def generate_test_case_excel(
    test_cases: list[dict[str, Any]],
    output_dir: str = "./excel_output",
    filename: Optional[str] = None,
) -> ExcelGenerationResult:
    """
    便捷函数：生成测试用例 Excel 文件.

    Args:
        test_cases: 测试用例列表
        output_dir: 输出目录
        filename: 输出文件名

    Returns:
        ExcelGenerationResult 生成结果
    """
    client = ExcelClient(output_dir=output_dir)
    return client.generate_excel(test_cases, filename)
