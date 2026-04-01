"""
Excel to Markdown Table Converter.

将 Excel (.xlsx/.xls) 文件转换为 Markdown 表格格式，供后续的 parse_mapping_node 使用。

支持功能:
1. 读取 .xlsx 和 .xls 文件
2. 提取所有工作表或指定工作表
3. 将每个工作表转换为 Markdown 表格
4. 合并多个表格为单个 Markdown 文档
"""

import logging
from pathlib import Path
from typing import Optional, Union

logger = logging.getLogger(__name__)


def excel_to_markdown(file_path: str, sheet_name: Optional[str] = None) -> str:
    """
    将 Excel 文件转换为 Markdown 表格。

    Args:
        file_path: Excel 文件路径
        sheet_name: 可选的工作表名称，如不提供则合并所有工作表

    Returns:
        Markdown 格式的表格字符串
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{file_path}")

    # 根据文件扩展名选择读取方法
    if path.suffix.lower() == '.xlsx':
        return _read_xlsx(path, sheet_name)
    elif path.suffix.lower() == '.xls':
        return _read_xls(path, sheet_name)
    else:
        raise ValueError(f"不支持的文件格式：{path.suffix}，仅支持 .xlsx 和 .xls")


def _read_xlsx(path: Path, sheet_name: Optional[str] = None) -> str:
    """读取 .xlsx 文件并转换为 Markdown 表格。"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("读取 .xlsx 文件需要 openpyxl 库：pip install openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # 获取工作表列表
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在，可用的工作表：{wb.sheetnames}")
        sheet_names_to_process = [sheet_name]
    else:
        sheet_names_to_process = wb.sheetnames

    logger.info(f"读取 Excel 文件：{path}，工作表：{sheet_names_to_process}")

    tables = []
    for name in sheet_names_to_process:
        ws = wb[name]
        table = _worksheet_to_markdown_table(ws, name)
        if table:
            tables.append(table)

    wb.close()

    return "\n\n".join(tables)


def _read_xls(path: Path, sheet_name: Optional[str] = None) -> str:
    """读取 .xls 文件并转换为 Markdown 表格。"""
    try:
        import xlrd
    except ImportError:
        raise ImportError("读取 .xls 文件需要 xlrd 库：pip install xlrd")

    wb = xlrd.open_workbook(path, formatting_info=False)

    # 获取工作表列表
    if sheet_name:
        if sheet_name not in wb.sheet_names():
            raise ValueError(f"工作表 '{sheet_name}' 不存在，可用的工作表：{wb.sheet_names()}")
        sheet_names_to_process = [sheet_name]
    else:
        sheet_names_to_process = wb.sheet_names()

    logger.info(f"读取 Excel 文件：{path}，工作表：{sheet_names_to_process}")

    tables = []
    for name in sheet_names_to_process:
        ws = wb.sheet_by_name(name)
        table = _xls_worksheet_to_markdown_table(ws, name)
        if table:
            tables.append(table)

    return "\n\n".join(tables)


def _worksheet_to_markdown_table(ws, sheet_name: str) -> str:
    """
    将 openpyxl 工作表转换为 Markdown 表格。

    Args:
        ws: openpyxl 工作表对象
        sheet_name: 工作表名称（用作标题）

    Returns:
        Markdown 表格字符串
    """
    # 获取所有有数据的行
    rows = []
    for row in ws.iter_rows(values_only=True):
        # 跳过全空的行
        if all(cell is None or (isinstance(cell, str) and cell.strip() == '') for cell in row):
            continue
        # 将 None 转换为空字符串
        rows.append([str(cell) if cell is not None else '' for cell in row])

    if not rows:
        logger.warning(f"工作表 '{sheet_name}' 没有数据")
        return ""

    return _rows_to_markdown_table(rows, sheet_name)


def _xls_worksheet_to_markdown_table(ws, sheet_name: str) -> str:
    """
    将 xlrd 工作表转换为 Markdown 表格。

    Args:
        ws: xlrd 工作表对象
        sheet_name: 工作表名称

    Returns:
        Markdown 表格字符串
    """
    rows = []
    for row_idx in range(ws.nrows):
        row = ws.row_values(row_idx)
        # 跳过全空的行
        if all(cell == '' or (isinstance(cell, str) and cell.strip() == '') for cell in row):
            continue
        rows.append([str(cell) if cell is not None else '' for cell in row])

    if not rows:
        logger.warning(f"工作表 '{sheet_name}' 没有数据")
        return ""

    return _rows_to_markdown_table(rows, sheet_name)


def _rows_to_markdown_table(rows: list[list[str]], sheet_name: str) -> str:
    """
    将行数据转换为 Markdown 表格。

    Args:
        rows: 行数据列表，每行是一个字符串列表
        sheet_name: 工作表名称（用作标题）

    Returns:
        Markdown 表格字符串
    """
    if not rows:
        return ""

    # 确定列数（以最长行为准）
    max_cols = max(len(row) for row in rows)

    # 补齐短行
    for row in rows:
        while len(row) < max_cols:
            row.append('')

    # 构建 Markdown 表格
    lines = []

    # 添加工作表标题（如果有多个工作表）
    lines.append(f"### {sheet_name}")
    lines.append("")

    # 表头
    header_row = rows[0]
    lines.append("| " + " | ".join(header_row) + " |")

    # 分隔线
    lines.append("|" + "|".join(["---"] * max_cols) + "|")

    # 数据行
    for row in rows[1:]:
        # 清理单元格内容（处理换行和管道符）
        cleaned_row = [cell.replace("\n", " ").replace("|", "\\|").strip() for cell in row]
        lines.append("| " + " | ".join(cleaned_row) + " |")

    return "\n".join(lines)


def convert_excel_to_markdown(file_path: Union[str, Path]) -> str:
    """
    便捷函数：将 Excel 文件转换为 Markdown 表格。

    Args:
        file_path: Excel 文件路径

    Returns:
        Markdown 格式的表格字符串
    """
    return excel_to_markdown(str(file_path))
